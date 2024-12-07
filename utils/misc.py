#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Sep 18 00:22:13 2020

@author: dl-asoro
"""
import sys
import os
import csv
import time
import functools
import shutil
import imp
import math
import pickle
import inspect
import logging
from openpyxl import load_workbook, Workbook
from importlib import reload
import torch
import numpy as np

def get_gpu_memory(acceptable_memory=256):
    import subprocess as sp
    
    free_memory = sp.check_output(['nvidia-smi', '--query-gpu=memory.free',
                                   '--format=csv,nounits,noheader'])
    total_memory = sp.check_output(['nvidia-smi', '--query-gpu=memory.total',
                                    '--format=csv,nounits,noheader'])
    
    free_memory = [int(x) for x in free_memory.decode('ascii').strip().split('\n')]
    total_memory = [int(x) for x in total_memory.decode('ascii').strip().split('\n')]
        
    # memory_map = dict(zip(range(len(free_memory)), free_memory))
    return free_memory, total_memory

def setup_gpu(gpu_id=0, memory=512, device_type='cuda', verbose=1):
    #setup gpu either by gpu_id or memory or both
    # os.environ["CUDA_DEVICE_ORDER"] = "PCI_BUS_ID"
    avail_memory, total_memory = get_gpu_memory()
    if verbose>0:
        for i in range(len(avail_memory)):
            print('GPU {} with Total:{} MiB, Free:{} MiB'.format(i, total_memory[i], avail_memory[i]))

    if verbose>0:
        print('Requested GPU Memory:', memory)
    # avail_memory = sorted(avail_memory)[::-1] # sort gpu by free memory
    # available_gpus = list(range(len(avail_memory)))
    available_gpus = [i for i, x in enumerate(avail_memory) if x>memory]
    if verbose>0:
        print('Available GPU:', available_gpus)
    # available_gpus = sorted(available_gpus)[::-1]
        
    if device_type=='cuda' and len(available_gpus)>0:
        if verbose>0:
            print('Using GPU:%d'% available_gpus[gpu_id])
        device = torch.device("cuda:{}".format(available_gpus[gpu_id]))
    else:
        if verbose>0:
            print('Using CPU')
        device = torch.device("cpu")
        
    return device

def create_new_dir(new_dir, clean=False):
    if clean:
        shutil.rmtree(new_dir, ignore_errors=True)
    if not os.path.exists(new_dir):
        os.makedirs(new_dir, exist_ok=True)
    return new_dir

def copy_src(root_src_dir, root_dst_dir, overwrite=True):
    for src_dir, dirs, files in os.walk(root_src_dir):
        dst_dir = src_dir.replace(root_src_dir, root_dst_dir, 1)
        if not os.path.exists(dst_dir):
            os.makedirs(dst_dir, exist_ok=True)
        for file in files:
            if 'cpython' in file:
                continue
            src_file = os.path.join(src_dir, file)
            dst_file = os.path.join(dst_dir, file)
            if os.path.exists(dst_file):
                if overwrite:
                    shutil.copy(src_file, dst_file)
            else:
                shutil.copy(src_file, dst_file)
                


def save_history(history, save_dir):
    # print('History saved to:', save_dir)
    row_head = list(history.keys())
    rows_val = np.around(np.array([val for key, val in history.items()]).T, 6)
    # row_num = rows_val[0].keys() #epoch number
    with open(save_dir+'/history.txt', 'a') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(row_head)
        writer.writerows(rows_val)
    csvfile.close()
    
def save_to_excel(file_path, row_data, header):
    # file_path = './out/GatedGCN/' + 'log_trial_results.xlsx'
    # Confirm file exists. 
    # If not, create it, add headers, then append new data
    try:
        wb = load_workbook(file_path)
        ws = wb.worksheets[0]  # select first worksheet
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.append(header) #header row
    ws.append(row_data)
    wb.save(file_path)
    wb.close()


def import_reload(module_name, *names):
    
    if module_name in sys.modules:
        reload(sys.modules[module_name])
    else:
        __import__(module_name, fromlist=names)

    for name in names:
        globals()[name] = getattr(sys.modules[module_name], name)

def create_logger(logPath=None):

    logging.basicConfig(format="%(asctime)s [%(levelname)-5.5s]  %(message)s", level=logging.INFO)
    rootLogger = logging.getLogger()

    # logFormatter = logging.Formatter("%(asctime)s [%(levelname)-5.5s]  %(message)s")

    # consoleHandler = logging.StreamHandler()
    # consoleHandler.setFormatter(logFormatter)
    # rootLogger.addHandler(consoleHandler)

    if logPath is not None:
        fileHandler = logging.FileHandler(os.path.join(logPath, 'logging.log'))
        # fileHandler.setFormatter(logFormatter)
        rootLogger.addHandler(fileHandler)

    return rootLogger